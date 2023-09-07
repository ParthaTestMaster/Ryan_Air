import inspect
import logging
from openpyxl import Workbook,load_workbook
import softest



class Utils(softest.TestCase):
    expected_url = 'https://www.amazon.co.uk/Nike-Park-Short-Sport-Shorts/dp/B07W5XYSSJ/ref=sr_1_10?crid=32RIA2KO1I72J&keywords=nike&qid=1686722397&sprefix=nike%2Caps%2C113&sr=8-10&th=1&psc=1'

    def assert_url(self):
        self.soft_assert(self.assertEqual, self.driver.current_url, self.expected_url)
        print('*******url assertion passed*******************')

    def cutom_logger(logLevel=logging.DEBUG):
        ##set class/ method name from where its called
        logger_name =  inspect.stack()[1][3]
        # # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel) ## always set lgging level to default
        # create console handler and File handler both
        # file handler D:\Ma$ter\pythonProject\pythonProject\amazon_nike\logs
        fh = logging.FileHandler("automation_logs.log", mode='w')
        # create formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s :%(message)s')
        # add formatter to ch and fh handler
        fh.setFormatter(formatter)
        # add ch to logger
        logger.addHandler(fh)
        return logger

    def read_data_from_excel(file_name_from_folder, sheet_name):
        data_list = []
        wb = load_workbook(filename=file_name_from_folder)
        sh = wb[sheet_name]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1, col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list
